"""Historical index values: daily basket/benchmark values used by the
webportal's own historic-return charts, populated manually or via Excel import."""
import io
from datetime import datetime

import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, UploadFile

from persistence import _load_historical_index, _save_historical_index

router = APIRouter()

@router.get("/api/index-history")
async def get_index_history():
    """Serve pre-computed historical index values for all baskets."""
    return _load_historical_index()


@router.post("/api/daily-values")
async def post_daily_values(body: dict):
    """Append (or update) daily basket + benchmark index values in historical_index.json.
    Body: { "date": "YYYY-MM-DD", "entries": [ { "basket": key, "value": float, "benchmark": float }, ... ] }
    If an entry for the given date already exists, it is overwritten."""
    date_str = (body.get("date") or "").strip()
    entries  = body.get("entries") or []

    if not date_str:
        raise HTTPException(status_code=400, detail="date is required")
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")

    hi = _load_historical_index()

    saved = []
    for entry in entries:
        basket = entry.get("basket", "").strip()
        value  = entry.get("value")
        bench  = entry.get("benchmark")
        if not basket or value is None or bench is None:
            continue
        if basket not in hi:
            continue
        data = hi[basket]["data"]
        # Remove existing entry for this date (overwrite)
        hi[basket]["data"] = [e for e in data if e["date"] != date_str]
        hi[basket]["data"].append({"date": date_str, "value": round(float(value), 4), "benchmark": round(float(bench), 4)})
        hi[basket]["data"].sort(key=lambda e: e["date"])
        saved.append(basket)

    _save_historical_index(hi)

    return {"ok": True, "date": date_str, "saved": saved}


@router.post("/api/import-excel-history")
async def import_excel_history(basket: str = Form(...), file: UploadFile = File(...)):
    """Import historical index values from an Excel file for a specific basket.
    Excel format: Column A = Date (YYYY-MM-DD), Column B = Basket Value, Column C = Benchmark.
    Only dates AFTER the last already-saved date are imported — existing data is never overwritten.
    """
    hi = _load_historical_index()

    if basket not in hi:
        raise HTTPException(status_code=400, detail=f"Unknown basket: {basket}")

    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    # Prefer a sheet with 'index' or 'value' in its name, else use first sheet
    sheet = next(
        (wb[n] for n in wb.sheetnames if any(k in n.lower() for k in ("index", "value", "historical"))),
        wb.active,
    )

    all_rows = list(sheet.iter_rows(values_only=True))
    if len(all_rows) < 2:
        raise HTTPException(status_code=400, detail="Excel file has no data rows")

    # Parse data rows — skip header (row 0); columns: 0=Date, 1=BasketValue, 2=Benchmark
    parsed = []
    for row in all_rows[1:]:
        if not row[0] or row[1] is None:
            continue
        date_raw = str(row[0]).strip().split(" ")[0]  # strip time component if present
        try:
            # Accept YYYY-MM-DD or DD-MM-YYYY
            if len(date_raw) == 10 and date_raw[4] == "-":
                date_str = date_raw
            else:
                from datetime import datetime as _dt
                date_str = _dt.strptime(date_raw, "%d-%m-%Y").strftime("%Y-%m-%d")
            datetime.strptime(date_str, "%Y-%m-%d")  # validate
        except Exception:
            continue
        try:
            value = round(float(str(row[1]).strip()), 4)
            benchmark = round(float(str(row[2]).strip()), 4) if row[2] is not None else None
        except Exception:
            continue
        if benchmark is None:
            continue
        parsed.append({"date": date_str, "value": value, "benchmark": benchmark})

    existing_dates = {e["date"] for e in hi[basket]["data"]}
    last_date = max(existing_dates) if existing_dates else "0000-00-00"

    # Only import dates strictly after the last saved date
    new_rows = [r for r in parsed if r["date"] > last_date]

    if not new_rows:
        return {"ok": True, "imported": 0, "lastDate": last_date,
                "message": f"Already up to date. Last saved date: {last_date}"}

    for row in new_rows:
        hi[basket]["data"] = [e for e in hi[basket]["data"] if e["date"] != row["date"]]
        hi[basket]["data"].append(row)
    hi[basket]["data"].sort(key=lambda e: e["date"])

    _save_historical_index(hi)

    return {
        "ok": True,
        "imported": len(new_rows),
        "lastDate": last_date,
        "newDates": [r["date"] for r in new_rows],
        "message": f"Imported {len(new_rows)} new date(s) after {last_date}",
    }


# ── Basket auto-detection from Excel column-B header ──────────────────────────
_BASKET_KEYWORDS = {
    "green energy":   "Green_Energy",
    "green":          "Green_Energy",
    "mid & small":    "Mid_Small_Cap",
    "mid and small":  "Mid_Small_Cap",
    "mid small":      "Mid_Small_Cap",
    "mid":            "Mid_Small_Cap",
    "ipo":            "IPO_Basket",
    "consumer trend": "Consumer_Trends",
    "consumer":       "Consumer_Trends",
    "trends trilogy": "Trends_Triology",
    "trends triology":"Trends_Triology",
    "triology":       "Trends_Triology",
    "trilogy":        "Trends_Triology",
    "techstack":      "Techstack",
    "tech stack":     "Techstack",
    "make in india":  "Make_in_India",
    "make":           "Make_in_India",
    "india":          "Make_in_India",
}

def _detect_basket(col_b_header: str) -> str | None:
    h = (col_b_header or "").lower()
    for kw, key in _BASKET_KEYWORDS.items():
        if kw in h:
            return key
    return None


def _parse_excel_rows(raw: bytes) -> tuple[list[dict], str]:
    """Return (parsed_rows, detected_basket_key). Raises ValueError on bad input."""
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = next(
        (wb[n] for n in wb.sheetnames if any(k in n.lower() for k in ("index", "value", "historical"))),
        wb.active,
    )
    all_rows = list(sheet.iter_rows(values_only=True))
    if len(all_rows) < 2:
        raise ValueError("Excel file has no data rows")

    header     = all_rows[0]
    basket_key = _detect_basket(str(header[1]) if len(header) > 1 else "")

    parsed = []
    for row in all_rows[1:]:
        if not row[0] or row[1] is None:
            continue
        date_raw = str(row[0]).strip().split(" ")[0]
        try:
            if len(date_raw) == 10 and date_raw[4] == "-":
                date_str = date_raw
            else:
                from datetime import datetime as _dt
                date_str = _dt.strptime(date_raw, "%d-%m-%Y").strftime("%Y-%m-%d")
            datetime.strptime(date_str, "%Y-%m-%d")
        except Exception:
            continue
        try:
            value     = round(float(str(row[1]).strip()), 4)
            benchmark = round(float(str(row[2]).strip()), 4) if len(row) > 2 and row[2] is not None else None
        except Exception:
            continue
        if benchmark is None:
            continue
        parsed.append({"date": date_str, "value": value, "benchmark": benchmark})

    return parsed, basket_key


@router.post("/api/import-excel-multi")
async def import_excel_multi(files: list[UploadFile] = File(...)):
    """Import multiple Excel files at once. Each file's basket is auto-detected
    from column B header. Only new dates (after the last saved entry) are added."""
    hi = _load_historical_index()

    results = []
    any_saved = False

    for upload in files:
        fname = upload.filename or "unknown"
        raw   = await upload.read()
        try:
            parsed, basket_key = _parse_excel_rows(raw)
        except Exception as e:
            results.append({"file": fname, "ok": False, "error": str(e)})
            continue

        if not basket_key or basket_key not in hi:
            results.append({"file": fname, "ok": False,
                            "error": f"Could not detect basket from column header. "
                                     f"Please rename column B to include the basket name (e.g. 'Green Energy Theme')."})
            continue

        existing_dates = {e["date"] for e in hi[basket_key]["data"]}
        last_date      = max(existing_dates) if existing_dates else "0000-00-00"
        new_rows       = [r for r in parsed if r["date"] > last_date]

        if not new_rows:
            results.append({"file": fname, "ok": True, "basket": basket_key,
                            "imported": 0, "lastDate": last_date,
                            "message": f"Already up to date (last saved: {last_date})"})
            continue

        for row in new_rows:
            hi[basket_key]["data"] = [e for e in hi[basket_key]["data"] if e["date"] != row["date"]]
            hi[basket_key]["data"].append(row)
        hi[basket_key]["data"].sort(key=lambda e: e["date"])
        any_saved = True

        results.append({"file": fname, "ok": True, "basket": basket_key,
                        "imported": len(new_rows), "lastDate": last_date,
                        "message": f"Imported {len(new_rows)} new date(s) after {last_date}"})

    if any_saved:
        _save_historical_index(hi)

    return {"ok": True, "results": results}


