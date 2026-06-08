"""
Process Mid & Small Cap rebalance history from Excel to generate buy_price_data.json entries.
Run once to rebuild Mid_Small_Cap buy/sell event data from scratch.
"""

import openpyxl
import json
import os
from collections import defaultdict
from datetime import datetime, timedelta

EXCEL_PATH = r"C:\Users\PC\OneDrive - Niveshaay Investment Advisors (1)\Nukul Madaan NIA\Niveshaay Office Work\Rebalance Historical Data\Rebalance Data.xlsx"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BUY_PRICE_DATA_PATH = os.path.join(BASE_DIR, "buy_price_data.json")
PORTFOLIOS_PATH = os.path.join(BASE_DIR, "portfolios.json")

MONTHS = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
          7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

CHANGE_THRESHOLD = 0.005  # ignore floating-point noise below this


def next_trading_day(dt: datetime) -> datetime:
    """Return next calendar day, skipping weekends."""
    d = dt + timedelta(days=1)
    while d.weekday() >= 5:  # 5=Sat, 6=Sun
        d += timedelta(days=1)
    return d


def fmt_date(dt: datetime) -> str:
    return f"{dt.day:02d} {MONTHS[dt.month]} {dt.year}"


def fmt_events(events):
    """Format list of (date_str, value) → newline-separated string (matches _parse_buy_events)."""
    return '\n'.join(f"{d} * {round(w, 4)}" for d, w in events)


def to_delta_events(buy_raw, sell_raw):
    """Convert total-weight buy events to delta weights (for prevBuyEvents legacy format).
    buy_raw: list of (date_str, total_weight); sell_raw: list of (date_str, delta_sold).
    Returns list of (date_str, delta_bought)."""
    from datetime import datetime as _dt
    import re as _re
    _MONTH_ABBR = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
                   'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
    def _ts(d):
        p = d.split()
        return _dt(int(p[2]), _MONTH_ABBR.get(p[1], 1), int(p[0]))

    combined = [(d, 'buy', q) for d, q in buy_raw] + [(d, 'sell', q) for d, q in sell_raw]
    combined.sort(key=lambda e: _ts(e[0]))
    cw = 0.0
    result = []
    for date_str, etype, qty in combined:
        if etype == 'buy':
            delta = qty - cw
            if delta > 0.001:
                result.append((date_str, round(delta, 4)))
            cw = qty
        else:
            cw = max(0.0, cw - qty)
    return result


def process_ticker(ticker, snapshots, sorted_dates):
    """
    Compare consecutive portfolio snapshots to derive buy/sell events.

    Absence from a snapshot means weight = 0 (stock removed at that date).

    buyEvents / sellEvents  = CURRENT series; stores TOTAL net weight per buy.
    prevBuyEvents / prevSellEvents = PREVIOUS series; stores DELTA weights (legacy format).
    sellEvents always stores DELTA sold.
    """
    s1_buy_total, s1_sell = [], []   # Series 1: total-weight buys + delta sells
    s2_buy_total, s2_sell = [], []   # Series 2: same

    prev_weight = 0.0
    had_reset = False

    for date in sorted_dates:
        new_weight = round(snapshots[date].get(ticker, 0.0), 6)

        delta = new_weight - prev_weight
        if abs(delta) < CHANGE_THRESHOLD:
            prev_weight = new_weight
            continue

        event_str = fmt_date(next_trading_day(date))

        if prev_weight < CHANGE_THRESHOLD and new_weight > CHANGE_THRESHOLD:
            # New addition
            if not had_reset:
                s1_buy_total.append((event_str, new_weight))
            else:
                s2_buy_total.append((event_str, new_weight))

        elif prev_weight > CHANGE_THRESHOLD and new_weight < CHANGE_THRESHOLD:
            # Full removal
            if not had_reset:
                s1_sell.append((event_str, round(prev_weight, 6)))
            else:
                s2_sell.append((event_str, round(prev_weight, 6)))
            had_reset = True

        elif delta > 0:
            # Weight increase → buy; store total new_weight
            if not had_reset:
                s1_buy_total.append((event_str, new_weight))
            else:
                s2_buy_total.append((event_str, new_weight))

        else:
            # Partial sell → store delta sold
            if not had_reset:
                s1_sell.append((event_str, round(-delta, 6)))
            else:
                s2_sell.append((event_str, round(-delta, 6)))

        prev_weight = new_weight

    if not had_reset:
        # Single series, never removed → goes to buyEvents (total-weight format)
        return {
            'buyEvents': fmt_events(s1_buy_total),
            'sellEvents': fmt_events(s1_sell),
            'prevBuyEvents': '',
            'prevSellEvents': '',
            'buyOHLC': {},
            'sellOHLC': {},
        }
    else:
        # prevBuyEvents uses legacy DELTA weights; buyEvents uses TOTAL weights
        s1_buy_delta = to_delta_events(s1_buy_total, s1_sell)
        return {
            'buyEvents': fmt_events(s2_buy_total),
            'sellEvents': fmt_events(s2_sell),
            'prevBuyEvents': fmt_events(s1_buy_delta),
            'prevSellEvents': fmt_events(s1_sell),
            'buyOHLC': {},
            'sellOHLC': {},
        }


def main():
    # 1. Read Excel
    print("Reading Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.worksheets[0]
    print(f"Sheet: {ws.title}, Rows: {ws.max_row}")

    rows = []
    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(r, 2).value
        ticker = ws.cell(r, 3).value
        weight = ws.cell(r, 6).value
        if ticker and date_val and weight is not None:
            rows.append((date_val, ticker.strip(), float(weight)))
    print(f"Loaded {len(rows)} rows")

    # 2. Build snapshot: date → {ticker: weight}
    snapshots = defaultdict(dict)
    for date, ticker, weight in rows:
        snapshots[date][ticker] = weight
    sorted_dates = sorted(snapshots.keys())
    print(f"Rebalance dates: {len(sorted_dates)} ({sorted_dates[0].date()} to {sorted_dates[-1].date()})")

    all_tickers = sorted(set(t for d, t, w in rows))
    print(f"Unique tickers: {len(all_tickers)}")

    # 3. Process each ticker using snapshot comparisons
    print("\nProcessing...")
    new_msc_data = {}
    for ticker in all_tickers:
        result = process_ticker(ticker, snapshots, sorted_dates)
        new_msc_data[ticker] = result

        has_prev = bool(result['prevBuyEvents'])
        has_cur  = bool(result['buyEvents'])
        label = "re-added" if has_prev and has_cur else ("sold(no re-add)" if has_prev else "active-single")
        print(f"  {ticker}: {label}")

    # 4. Update buy_price_data.json
    print(f"\nUpdating {BUY_PRICE_DATA_PATH}...")
    with open(BUY_PRICE_DATA_PATH, 'r') as f:
        buy_price_data = json.load(f)
    buy_price_data['Mid_Small_Cap'] = new_msc_data
    with open(BUY_PRICE_DATA_PATH, 'w') as f:
        json.dump(buy_price_data, f, indent=2)
    print("buy_price_data.json updated.")

    # 5. Active stocks at last rebalance
    last_date = sorted_dates[-1]
    active = {t: w for t, w in snapshots[last_date].items()}
    print(f"\nActive at {last_date.date()} ({len(active)} stocks):")
    for t, w in sorted(active.items()):
        print(f"  {t}: {round(w, 4)}%")

    sold = set(all_tickers) - set(active.keys())
    print(f"\nSold/removed ({len(sold)} stocks):")
    print(sorted(sold))

    print("\nDone.")


if __name__ == '__main__':
    main()
