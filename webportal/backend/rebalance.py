"""Rebalance workflow: trigger price refresh, upload rebalance Excel, preview
the computed diff, confirm and apply it, and the activity log endpoint.
This is the largest and most stateful part of the app -- it touches
portfolios, buy-price data, rebalance history, and gains all at once, and
auto-snapshots for rollback/undo before any destructive write.
"""
import csv
import io
import json
import re
import time
from datetime import date as _date, datetime, timezone

import openpyxl
from fastapi import APIRouter, BackgroundTasks, File, Form, HTTPException, Request, UploadFile

import price_engine
from buy_price_gains import (
    _date_to_ts, _parse_buy_events, _add_event,
    _rebuild_sold_from_bp, _recalc_basket_buy_prices,
    _refresh_gains_file, _backfill_all_sell_ohlc_bg,
)
from persistence import (
    BASKET_DISPLAY_NAMES, _ACTIVITY_LOG_FILE,
    _auto_save_rollback, _push_undo_snapshot, _require_admin, _log_activity,
    _load_portfolios, _save_portfolios,
    _load_buy_price_data, _save_buy_price_data,
    _load_rebalance_history, _save_rebalance_history,
)
from price_engine import _norm_name, _resolve_nse, _fetch_nse_symbols
from common.admin import ADMIN_EMAILS
from live_data import _fetch_rebalance_prices

router = APIRouter()

@router.post("/api/trigger-rebalance")
async def trigger_rebalance(background_tasks: BackgroundTasks, basket: str = Form(...)):
    if basket not in BASKET_DISPLAY_NAMES:
        raise HTTPException(status_code=400, detail=f"Unknown basket: {basket}")
    background_tasks.add_task(_recalc_basket_buy_prices, basket)
    background_tasks.add_task(_refresh_gains_file)
    return {"ok": True, "message": f"Rebalance triggered for {BASKET_DISPLAY_NAMES[basket]}"}


_DATE_RE = re.compile(
    r'\b(\d{1,2})[\/\-\s](\w{3,9})[\/\-\s](\d{4})\b'
    r'|\b(\d{4})[\/\-](\d{1,2})[\/\-](\d{1,2})\b'
    r'|\b(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})\b',
    re.IGNORECASE,
)
_MON = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
        "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12,
        "january":1,"february":2,"march":3,"april":4,"june":6,
        "july":7,"august":8,"september":9,"october":10,"november":11,"december":12}

def _parse_date_value(val) -> str | None:
    """Convert any cell value / string to 'DD Mon YYYY', or None."""
    if val is None:
        return None
    if isinstance(val, (datetime, _date)):
        try:
            return val.strftime("%d %b %Y")
        except Exception:
            return None
    s = str(val).strip()
    # try standard strptime formats first
    for fmt in ("%d %b %Y", "%d-%b-%Y", "%d/%m/%Y", "%d-%m-%Y",
                "%Y-%m-%d", "%d %B %Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d %b %Y")
        except ValueError:
            pass
    # regex scan — picks up dates embedded in strings like "Rebalance as on 15 May 2026"
    for m in _DATE_RE.finditer(s):
        g = m.groups()
        try:
            if g[0]:   # DD Mon YYYY
                mon = _MON.get(g[1].lower()[:3])
                if mon:
                    return datetime(int(g[2]), mon, int(g[0])).strftime("%d %b %Y")
            elif g[3]: # YYYY-MM-DD
                return datetime(int(g[3]), int(g[4]), int(g[5])).strftime("%d %b %Y")
            elif g[6]: # DD-MM-YYYY
                return datetime(int(g[8]), int(g[7]), int(g[6])).strftime("%d %b %Y")
        except ValueError:
            continue
    return None


def _extract_rebalance_date(wb, filename: str = "") -> str | None:
    """Find the rebalance date from an openpyxl workbook: sheet name → first 6 rows → filename."""
    # 1. Sheet name
    for name in wb.sheetnames:
        d = _parse_date_value(name)
        if d:
            return d
    # 2. First 6 rows of first sheet
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=1, max_row=6, values_only=True):
        for cell in row:
            d = _parse_date_value(cell)
            if d:
                return d
    # 3. Filename
    return _parse_date_value(filename)


_REBALANCE_ALLOWED = ADMIN_EMAILS

@router.post("/api/upload-rebalance")
async def upload_rebalance(
    request: Request,
    background_tasks: BackgroundTasks,
    basket: str = Form(...),
    file: UploadFile = File(...),
):
    user_email = request.headers.get("X-User-Email", "")
    if user_email and user_email not in _REBALANCE_ALLOWED:
        raise HTTPException(status_code=403, detail="You do not have permission to upload rebalance files.")
    if basket not in BASKET_DISPLAY_NAMES:
        raise HTTPException(status_code=400, detail=f"Unknown basket: {basket}")

    raw = await file.read()
    fname = (file.filename or "").lower()

    date_str: str | None = None
    new_stocks: list = []

    if fname.endswith((".xlsx", ".xls")):
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)

        # Target "Historical Constituents" sheet; fall back to sheet 2, then sheet 1
        const_sheet = next(
            (ws for ws in wb.worksheets if "constituent" in (ws.title or "").lower()),
            wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0]
        )

        all_rows = list(const_sheet.iter_rows(values_only=True))
        wb.close()

        if not all_rows:
            raise HTTPException(400, "Historical Constituents sheet is empty.")

        # Find header row (contains 'date', 'constituent', or 'weight')
        hdr_idx = 0
        for i, row in enumerate(all_rows[:6]):
            if any("date" in str(c).lower() or "constituent" in str(c).lower() or "weight" in str(c).lower()
                   for c in row if c is not None):
                hdr_idx = i
                break

        headers = [str(c).strip().lower() if c is not None else "" for c in all_rows[hdr_idx]]

        # Column indices: A=date, B=constituent, C=weight (by header keyword, fallback to position)
        date_col   = next((i for i, h in enumerate(headers) if "date" in h), 0)
        const_col  = next((i for i, h in enumerate(headers)
                           if any(k in h for k in ("constituent", "nse", "symbol", "ticker"))), 1)
        weight_col = next((i for i, h in enumerate(headers) if "weight" in h), 2)

        # Parse rows; carry forward last seen date to handle merged cells
        date_buckets: dict[str, list] = {}
        cur_date: str | None = None
        for row in all_rows[hdr_idx + 1:]:
            if all(c is None for c in row):
                continue

            date_val = row[date_col] if len(row) > date_col else None
            if date_val is not None:
                # Date Range format: 'YYYY-MM-DD to YYYY-MM-DD' — split on ' to ', take start date
                raw_s = str(date_val).strip()
                parts = re.split(r'\s+to\s+', raw_s, flags=re.IGNORECASE)
                d = _parse_excel_date(parts[0].strip()) or _parse_date_value(parts[0].strip())
                if d:
                    cur_date = d

            if not cur_date:
                continue

            const_val  = row[const_col]  if len(row) > const_col  else None
            weight_val = row[weight_col] if len(row) > weight_col else None

            # Column B is the company name (full name like "Zen Technologies Ltd")
            name = str(const_val).strip() if const_val is not None else ""
            if not name or name.lower() in ("constituents", "nse code", "symbol", "ticker", "name", "none"):
                continue

            try:
                weight = float(str(weight_val).strip().rstrip("%")) if weight_val is not None else 0.0
            except (ValueError, AttributeError):
                continue
            if weight <= 0:
                continue

            date_buckets.setdefault(cur_date, []).append((name, weight))

        if not date_buckets:
            raise HTTPException(400, "No valid stock data found in Historical Constituents sheet. "
                                     "Expected columns: Date Range, Constituents, Weightage")

        # --- Multi-date processing: all dates chronologically with their original dates ---
        all_dates_sorted = sorted(date_buckets.keys(), key=lambda d: _date_to_ts(d))

        rh_pre         = _load_rebalance_history()
        existing_dates = {e.get("date", "").strip() for e in rh_pre.get(basket, [])}
        new_dates      = [d for d in all_dates_sorted if d not in existing_dates]

        if not new_dates:
            return {"duplicate": True, "message": "File has been already uploaded"}

        pf_data       = _load_portfolios()
        curr_stocks_l = pf_data.get(basket, [])
        nse_sym_list  = price_engine._nse_symbols_cache or await _fetch_nse_symbols()

        # Build name → NSE code reverse map.
        # Sources (in priority order): buy_price_data securityName, then rebalance history.
        # Shorter codes always win to avoid full-name fallbacks (e.g. "SJS" beats "SJS ENTERPRISES LTD").
        def _build_name_map(pairs: list[tuple[str, str]]) -> dict[str, str]:
            m: dict[str, str] = {}
            for sn, cd in pairs:
                if sn and cd:
                    k = _norm_name(sn)
                    if not m.get(k) or len(cd) < len(m[k]):
                        m[k] = cd
            return m

        bp_lookup   = _load_buy_price_data().get(basket, {})
        rh_lookup   = _load_rebalance_history().get(basket, [])
        history_name_map: dict[str, str] = _build_name_map(
            [(det.get("securityName", ""), code) for code, det in bp_lookup.items()]
            + [(e.get("securityName", ""), e.get("nseCode", "")) for e in rh_lookup]
        )

        # Build resolved stock list per new date
        date_stock_map: dict[str, list] = {}
        for d in new_dates:
            raw_entries = date_buckets[d]
            weight_sum  = sum(w for _, w in raw_entries)
            scale       = 100.0 if weight_sum <= 2.0 else 1.0
            stocks_d    = []
            for name, weight in raw_entries:
                nse = (history_name_map.get(_norm_name(name))
                       or _resolve_nse(name, curr_stocks_l, nse_sym_list)
                       or name.upper())
                stocks_d.append({
                    "nseCode": nse, "securityName": name, "segment": "Equity",
                    "weight": round(weight * scale, 4), "date": d,
                })
            if stocks_d:
                date_stock_map[d] = stocks_d

        if not date_stock_map:
            raise HTTPException(400, "No valid stocks found after NSE resolution.")

        _auto_save_rollback()
        _push_undo_snapshot(basket, f"before rebalance {new_dates[-1]}")

        # Only track stocks present in the LATEST date range (current portfolio).
        # Stocks that existed in older ranges but not the latest are completely ignored.
        latest_new = new_dates[-1]
        current_codes = {s["nseCode"] for s in date_stock_map.get(latest_new, [])}

        # Per-date lookup: code → stock entry
        date_snaps = {d: {s["nseCode"]: s for s in date_stock_map[d]}
                      for d in new_dates if d in date_stock_map}

        # Load existing history to know prior weights for each current-portfolio stock
        rh          = _load_rebalance_history()
        bh          = rh.get(basket, [])
        by_date_h: dict = {}
        for e in bh:
            by_date_h.setdefault(e.get("date", ""), []).append(e)
        latest_existing = max(by_date_h, key=lambda d: _date_to_ts(d), default=None)
        existing_weights = (
            {e["nseCode"]: float(e.get("weight", 0))
             for e in by_date_h.get(latest_existing, [])}
            if latest_existing else {}
        )

        portfolios  = _load_portfolios()
        basket_stks = portfolios.get(basket, [])
        stk_map     = {s["nseCode"]: s for s in basket_stks}
        bp_data     = _load_buy_price_data()
        basket_bp   = bp_data.setdefault(basket, {})

        all_summary_rows: list[dict]   = []
        bg_added_per_date: dict        = {d: [] for d in new_dates}
        bg_sold_per_date:  dict        = {d: [] for d in new_dates}  # sell codes per date

        # ── Stock-centric loop: trace each current-portfolio stock through history ──
        for code in current_codes:
            prev_w = existing_weights.get(code, 0.0)

            for cur_date in new_dates:
                day_snap = date_snaps.get(cur_date, {})
                if code not in day_snap:
                    # Stock absent from this date range; reset so next appearance = fresh buy
                    prev_w = 0.0
                    continue

                s      = day_snap[code]
                new_w  = s["weight"]
                det_sn = s.get("securityName", "")
                det_sg = s.get("segment", "Equity")

                # If prev_w was reset to 0 because the stock was absent from an
                # intermediate Excel date range (not a true first entry), recover
                # the actual prior weight from rebalance history so we store the
                # delta instead of the full cumulative weight.
                if prev_w == 0 and code in basket_bp and basket_bp[code].get("buyEvents"):
                    det_bp = basket_bp[code]
                    # Respect series boundary: only look at entries AFTER the last
                    # prevSellEvents date (so true re-entries after full exits still
                    # get treated as first buys in the new series).
                    prev_sell_lines = [
                        ln.strip().split(" * ")[0].strip()
                        for ln in (det_bp.get("prevSellEvents") or "").strip().split("\n")
                        if " * " in ln.strip()
                    ]
                    boundary_ts = max((_date_to_ts(d) for d in prev_sell_lines), default=0)
                    cur_ts      = _date_to_ts(cur_date)
                    recovered_w = max(
                        (float(e.get("weight", 0))
                         for e in bh
                         if e.get("nseCode") == code
                         and boundary_ts < _date_to_ts(e.get("date", "")) < cur_ts),
                        default=0.0,
                    )
                    if recovered_w > 0:
                        prev_w = recovered_w

                if prev_w == 0:
                    # True first appearance (new stock, no prior history in this series)
                    _add_event(basket_bp, code, "buyEvents", cur_date, new_w)
                    det = basket_bp[code]
                    if not det.get("securityName"):
                        det["securityName"] = det_sn
                    if not det.get("segment"):
                        det["segment"] = det_sg
                    all_summary_rows.append({
                        "nseCode": code, "securityName": det_sn,
                        "date": cur_date, "prevWeight": 0,
                        "newWeight": round(new_w, 2), "action": "Added",
                    })
                    bg_added_per_date[cur_date].append(code)

                elif new_w > prev_w + 0.01:
                    delta = round(new_w - prev_w, 4)
                    _add_event(basket_bp, code, "buyEvents", cur_date, delta)
                    all_summary_rows.append({
                        "nseCode": code, "securityName": det_sn,
                        "date": cur_date, "prevWeight": round(prev_w, 2),
                        "newWeight": round(new_w, 2), "action": "Increased",
                    })
                    bg_added_per_date[cur_date].append(code)

                elif new_w < prev_w - 0.01:
                    delta = round(prev_w - new_w, 4)
                    _add_event(basket_bp, code, "sellEvents", cur_date, delta)
                    all_summary_rows.append({
                        "nseCode": code, "securityName": det_sn,
                        "date": cur_date, "prevWeight": round(prev_w, 2),
                        "newWeight": round(new_w, 2), "action": "Decreased",
                    })
                    bg_sold_per_date[cur_date].append(code)

                prev_w = new_w

            # Ensure portfolio entry reflects current allocation
            latest_s = date_snaps.get(latest_new, {}).get(code)
            if latest_s:
                alloc = round(latest_s["weight"] / 100, 6)
                if code in stk_map:
                    stk_map[code]["allocation"] = alloc
                else:
                    entry = {"nseCode": code, "allocation": alloc, "buyPrice": None}
                    basket_stks.append(entry)
                    stk_map[code] = entry

        # Add to rebalance history — save ALL stocks at every new date so the date
        # is always tracked in existing_dates on the next upload.
        for d in new_dates:
            for s in date_stock_map.get(d, []):
                rh.setdefault(basket, []).append(s)

        # Record full exits: stocks that had weight before this upload but are no longer
        # in the portfolio (not in current_codes). Write a sell event for their remaining weight.
        for code, prev_weight in existing_weights.items():
            if prev_weight <= 0 or code in current_codes:
                continue  # still active or was already at 0
            det = basket_bp.get(code)
            if det is None:
                continue
            all_buys  = _parse_buy_events(det.get("buyEvents")  or "")
            all_sells = _parse_buy_events(det.get("sellEvents") or "")
            total_bought = sum(q for _, q in all_buys)
            total_sold   = sum(q for _, q in all_sells)
            remaining    = round(total_bought - total_sold, 4)
            if remaining > 0.01:
                _add_event(basket_bp, code, "sellEvents", latest_new, remaining)
                bg_sold_per_date[latest_new].append(code)
                all_summary_rows.append({
                    "nseCode": code,
                    "securityName": det.get("securityName", ""),
                    "date": latest_new, "prevWeight": round(prev_weight, 2),
                    "newWeight": 0, "action": "Removed",
                })

        # Persist
        _save_rebalance_history(rh)
        portfolios[basket] = basket_stks
        # Rebuild sold list from the updated buy/sell events so new sell events
        # appear in the P&L / Sold Stocks tab immediately after upload.
        old_sold = portfolios.get(f"{basket}_sold", [])
        portfolios[f"{basket}_sold"] = _rebuild_sold_from_bp(basket_bp, old_sold)
        _save_portfolios(portfolios)
        _save_buy_price_data(bp_data)

        for d in new_dates:
            has_buys  = bool(bg_added_per_date.get(d))
            has_sells = bool(bg_sold_per_date.get(d))
            if has_buys or has_sells:
                background_tasks.add_task(_fetch_rebalance_prices, basket, d,
                                          bg_added_per_date.get(d, []),
                                          bg_sold_per_date.get(d, []))
        background_tasks.add_task(_recalc_basket_buy_prices, basket)
        background_tasks.add_task(_refresh_gains_file)

        return {
            "ok": True,
            "basket": BASKET_DISPLAY_NAMES[basket],
            "date": latest_new,
            "datesProcessed": new_dates,
            "stockCount": len(current_codes),
            "summary": all_summary_rows,
        }

    else:
        # CSV fallback: date from filename
        date_str = _parse_date_value(file.filename or "")
        if not date_str:
            raise HTTPException(400, "Could not find rebalance date. Include the date in the CSV filename.")

        def _parse_rebalance_row(row: dict) -> dict | None:
            nse = (row.get("NSE Code") or row.get("nseCode") or row.get("NSE") or
                   row.get("Symbol") or row.get("Ticker") or "")
            nse = nse.strip().upper() if isinstance(nse, str) else str(nse).strip().upper()
            if not nse or nse in ("NONE", "NSE CODE", "TICKER", "SYMBOL"):
                return None
            name = str(row.get("Security Name") or row.get("Name") or "").strip()
            w_raw = (row.get("Weightage (%)") or row.get("Weightage") or row.get("Weight (%)") or
                     row.get("Weight") or row.get("weight") or row.get("Allocation (%)") or 0)
            try:
                weight = float(str(w_raw).strip().rstrip("%"))
            except (ValueError, AttributeError):
                return None
            return {"nseCode": nse, "securityName": name, "segment": "Equity", "weight": weight}

        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            parsed = _parse_rebalance_row(dict(row))
            if parsed:
                parsed["date"] = date_str
                new_stocks.append(parsed)

    # Duplicate check (after date is known)
    rh = _load_rebalance_history()
    existing_dates = {e.get("date", "").strip() for e in rh.get(basket, [])}
    if date_str in existing_dates:
        return {"duplicate": True, "message": "File has been already uploaded"}

    # Snapshot before applying rebalance changes
    _auto_save_rollback()
    _push_undo_snapshot(basket, f"before rebalance {date_str}")

    if not new_stocks:
        raise HTTPException(status_code=400, detail="No valid stocks found. "
                            "Expected columns: NSE Code, Security Name, Weightage (%), Segment")

    # Latest snapshot for diff
    basket_history = rh.get(basket, [])
    by_date: dict = {}
    for e in basket_history:
        by_date.setdefault(e.get("date", ""), []).append(e)

    latest_date = max(by_date, key=lambda d: _date_to_ts(d), default=None)
    prev_snap   = {e["nseCode"]: e for e in by_date.get(latest_date, [])} if latest_date else {}
    new_snap    = {s["nseCode"]: s for s in new_stocks}

    # Compute diff
    added     = [c for c in new_snap if c not in prev_snap]
    removed   = [c for c in prev_snap if c not in new_snap]
    increased = []
    decreased = []
    for code in new_snap:
        if code in prev_snap:
            old_w = float(prev_snap[code].get("weight", 0))
            new_w = new_snap[code]["weight"]
            if new_w > old_w + 0.01:
                increased.append({"nseCode": code, "from": old_w, "to": new_w})
            elif new_w < old_w - 0.01:
                decreased.append({"nseCode": code, "from": old_w, "to": new_w})

    # 1. Persist to rebalance_history.json
    rh.setdefault(basket, []).extend(new_stocks)
    _save_rebalance_history(rh)

    # 2. Update portfolios.json
    portfolios = _load_portfolios()
    basket_stocks = portfolios.get(basket, [])
    stk_map = {s["nseCode"]: s for s in basket_stocks}
    sold = portfolios.get(f"{basket}_sold", [])

    # New stocks
    for code in added:
        s = new_snap[code]
        if code in stk_map:
            stk_map[code]["allocation"] = round(s["weight"] / 100, 6)
        else:
            entry = {"nseCode": code, "allocation": round(s["weight"] / 100, 6), "buyPrice": None}
            basket_stocks.append(entry)
            stk_map[code] = entry

    # Weight changes for existing stocks
    for item in increased + decreased:
        code = item["nseCode"]
        if code in stk_map:
            stk_map[code]["allocation"] = round(new_snap[code]["weight"] / 100, 6)

    # Wholly removed → sell event (buyPrice computed in background via FIFO/wavg)
    for code in removed:
        old = prev_snap[code]
        sold.append({
            "nseCode": code,
            "securityName": old.get("securityName", ""),
            "date": date_str,
            "action": "Wholly Sold",
            "weightSold": round(float(old.get("weight", 0)), 2),
            "buyPrice": None,
            "sellPrice": None,
        })
        basket_stocks = [s for s in basket_stocks if s["nseCode"] != code]
        stk_map.pop(code, None)

    # Partial reductions → sell event
    for item in decreased:
        code = item["nseCode"]
        old = prev_snap[code]
        sold.append({
            "nseCode": code,
            "securityName": old.get("securityName", ""),
            "date": date_str,
            "action": "Partial Sell",
            "weightSold": round(item["from"] - item["to"], 2),
            "buyPrice": None,
            "sellPrice": None,
        })

    portfolios[basket] = basket_stocks
    portfolios[f"{basket}_sold"] = sold
    _save_portfolios(portfolios)

    # 3b. Update buy_price_data.json with buy/sell event text entries
    bp_data   = _load_buy_price_data()
    basket_bp = bp_data.setdefault(basket, {})

    for code in added:
        s = new_snap[code]
        _add_event(basket_bp, code, "buyEvents", date_str, s["weight"])
        det = basket_bp[code]
        if not det.get("securityName"):
            det["securityName"] = s.get("securityName", "")
        if not det.get("segment"):
            det["segment"] = s.get("segment", "Equity")

    for item in increased:
        delta = round(item["to"] - item["from"], 4)
        _add_event(basket_bp, item["nseCode"], "buyEvents", date_str, delta)

    for code in removed:
        old_w = float(prev_snap[code].get("weight", 0))
        _add_event(basket_bp, code, "sellEvents", date_str, old_w)

    for item in decreased:
        delta = round(item["from"] - item["to"], 4)
        _add_event(basket_bp, item["nseCode"], "sellEvents", date_str, delta)

    _save_buy_price_data(bp_data)

    # 4. Background: fetch OHLC prices, recalc buy prices, refresh gains
    sell_codes = removed + [i["nseCode"] for i in decreased]
    background_tasks.add_task(_fetch_rebalance_prices, basket, date_str, added, sell_codes)
    background_tasks.add_task(_recalc_basket_buy_prices, basket)
    background_tasks.add_task(_refresh_gains_file)

    # Build structured summary (only changed stocks, weights in % form)
    summary_rows = []
    for code in added:
        s = new_snap[code]
        summary_rows.append({"nseCode": code, "securityName": s.get("securityName", ""),
                              "prevWeight": 0, "newWeight": round(float(s["weight"]), 2), "action": "Added"})
    for code in removed:
        old = prev_snap[code]
        prev_w = float(old.get("weight", 0))
        summary_rows.append({"nseCode": code, "securityName": old.get("securityName", ""),
                              "prevWeight": round(prev_w, 2), "newWeight": 0, "action": "Removed"})
    for item in increased:
        code = item["nseCode"]
        summary_rows.append({"nseCode": code, "securityName": new_snap[code].get("securityName", ""),
                              "prevWeight": round(item["from"], 2), "newWeight": round(item["to"], 2), "action": "Increased"})
    for item in decreased:
        code = item["nseCode"]
        summary_rows.append({"nseCode": code, "securityName": new_snap[code].get("securityName", ""),
                              "prevWeight": round(item["from"], 2), "newWeight": round(item["to"], 2), "action": "Decreased"})

    return {
        "ok": True,
        "basket": BASKET_DISPLAY_NAMES[basket],
        "date": date_str,
        "stockCount": len(new_stocks),
        "summary": summary_rows,
    }


@router.post("/api/preview-rebalance")
async def preview_rebalance(
    request: Request,
    basket: str = Form(...),
    file: UploadFile = File(...),
):
    """Parse Excel Sheet 2 and return slide1 + slide2 preview WITHOUT writing anything."""
    _require_admin(request)
    if basket not in BASKET_DISPLAY_NAMES:
        raise HTTPException(status_code=400, detail=f"Unknown basket: {basket}")

    raw = await file.read()
    fname = (file.filename or "").lower()
    if not fname.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls files are supported for rebalance preview.")

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    const_sheet = next(
        (ws for ws in wb.worksheets if "constituent" in (ws.title or "").lower()),
        wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0],
    )
    all_rows = list(const_sheet.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise HTTPException(400, "Historical Constituents sheet is empty.")

    hdr_idx = 0
    for i, row in enumerate(all_rows[:6]):
        if any("date" in str(c).lower() or "constituent" in str(c).lower() or "weight" in str(c).lower()
               for c in row if c is not None):
            hdr_idx = i
            break

    headers    = [str(c).strip().lower() if c is not None else "" for c in all_rows[hdr_idx]]
    date_col   = next((i for i, h in enumerate(headers) if "date" in h), 0)
    const_col  = next((i for i, h in enumerate(headers)
                       if any(k in h for k in ("constituent", "nse", "symbol", "ticker"))), 1)
    weight_col = next((i for i, h in enumerate(headers) if "weight" in h), 2)

    date_buckets: dict[str, list] = {}
    cur_date: str | None = None
    for row in all_rows[hdr_idx + 1:]:
        if all(c is None for c in row):
            continue
        date_val = row[date_col] if len(row) > date_col else None
        if date_val is not None:
            raw_s = str(date_val).strip()
            parts = re.split(r'\s+to\s+', raw_s, flags=re.IGNORECASE)
            d = _parse_excel_date(parts[0].strip()) or _parse_date_value(parts[0].strip())
            if d:
                cur_date = d
        if not cur_date:
            continue
        const_val  = row[const_col]  if len(row) > const_col  else None
        weight_val = row[weight_col] if len(row) > weight_col else None
        name = str(const_val).strip() if const_val is not None else ""
        if not name or name.lower() in ("constituents", "nse code", "symbol", "ticker", "name", "none"):
            continue
        try:
            weight = float(str(weight_val).strip().rstrip("%")) if weight_val is not None else 0.0
        except (ValueError, AttributeError):
            continue
        if weight <= 0:
            continue
        date_buckets.setdefault(cur_date, []).append((name, weight))

    if not date_buckets:
        raise HTTPException(400, "No valid stock data found in Historical Constituents sheet.")

    all_dates_sorted = sorted(date_buckets.keys(), key=lambda d: _date_to_ts(d))
    rh_pre         = _load_rebalance_history()
    existing_dates = {e.get("date", "").strip() for e in rh_pre.get(basket, [])}
    new_dates      = [d for d in all_dates_sorted if d not in existing_dates]

    if not new_dates:
        latest_in_file = all_dates_sorted[-1] if all_dates_sorted else "unknown"
        latest_existing = max(existing_dates, key=lambda d: _date_to_ts(d)) if existing_dates else "none"
        return {
            "duplicate": True,
            "message": (
                f"No new dates found in this file. "
                f"Latest date detected in file: {latest_in_file}. "
                f"Latest date already in system: {latest_existing}. "
                f"Please ensure the new rebalance date has been added to the Excel file before uploading."
            ),
        }

    # Build name → NSE code reverse map
    bp_lookup = _load_buy_price_data().get(basket, {})
    rh_lookup = rh_pre.get(basket, [])
    history_name_map: dict[str, str] = {}
    for code, det in bp_lookup.items():
        sn = det.get("securityName", "")
        if sn and code:
            k = _norm_name(sn)
            if not history_name_map.get(k) or len(code) < len(history_name_map[k]):
                history_name_map[k] = code
    for e in rh_lookup:
        sn, code = e.get("securityName", ""), e.get("nseCode", "")
        if sn and code:
            k = _norm_name(sn)
            if not history_name_map.get(k) or len(code) < len(history_name_map[k]):
                history_name_map[k] = code

    pf_data       = _load_portfolios()
    curr_stocks_l = pf_data.get(basket, [])
    nse_sym_list  = price_engine._nse_symbols_cache or await _fetch_nse_symbols()

    date_stock_map: dict[str, list] = {}
    for d in new_dates:
        raw_entries = date_buckets[d]
        weight_sum  = sum(w for _, w in raw_entries)
        scale       = 100.0 if weight_sum <= 2.0 else 1.0
        stocks_d: list = []
        for name, weight in raw_entries:
            nse = (history_name_map.get(_norm_name(name))
                   or _resolve_nse(name, curr_stocks_l, nse_sym_list)
                   or name.upper())
            stocks_d.append({
                "nseCode": nse, "securityName": name, "segment": "Equity",
                "weight": round(weight * scale, 4), "date": d,
            })
        if stocks_d:
            date_stock_map[d] = stocks_d

    if not date_stock_map:
        raise HTTPException(400, "No valid stocks found after name resolution.")

    latest_new    = new_dates[-1]
    current_codes = {s["nseCode"] for s in date_stock_map.get(latest_new, [])}
    date_snaps    = {d: {s["nseCode"]: s for s in date_stock_map[d]}
                     for d in new_dates if d in date_stock_map}

    bh = rh_pre.get(basket, [])
    by_date_h: dict = {}
    for e in bh:
        by_date_h.setdefault(e.get("date", ""), []).append(e)
    latest_existing = max(by_date_h, key=lambda d: _date_to_ts(d), default=None)
    existing_weights = (
        {e["nseCode"]: float(e.get("weight", 0)) for e in by_date_h.get(latest_existing, [])}
        if latest_existing else {}
    )

    # Supplement existing_weights with current portfolio holdings.
    # Old rebalance_history entries may be incomplete (written with a prior code bug that
    # filtered by current_codes). Any stock currently in the portfolio but absent from
    # rebalance_history would be misclassified as "New Addition" on the next upload,
    # generating a duplicate buy event. The live portfolio is the authoritative baseline.
    for s in curr_stocks_l:
        code  = s["nseCode"]
        alloc = s.get("allocation") or 0
        if code not in existing_weights and alloc > 0:
            existing_weights[code] = round(float(alloc) * 100, 4)

    basket_bp_curr = _load_buy_price_data().get(basket, {})

    # Only generate buy/sell events for dates that come AFTER the latest existing baseline.
    # New dates that are chronologically earlier than latest_existing are "historical gap"
    # dates (present in the Excel but missing from the DB). Processing them from the
    # latest_existing weights would incorrectly flag long-absent stocks as exits/re-entries.
    # They are still saved to rebalance_history (so they won't be reprocessed next time)
    # but no new events are emitted for them.
    if latest_existing:
        event_dates = [d for d in new_dates if _date_to_ts(d) > _date_to_ts(latest_existing)]
    else:
        event_dates = list(new_dates)

    codes_to_process: set[str] = set(existing_weights.keys())
    for d in event_dates:
        for s in date_stock_map.get(d, []):
            codes_to_process.add(s["nseCode"])

    historical_events: list[dict] = []
    latest_events:     list[dict] = []

    for code in codes_to_process:
        prev_w   = existing_weights.get(code, 0.0)
        last_sn  = basket_bp_curr.get(code, {}).get("securityName", code)
        last_seg = basket_bp_curr.get(code, {}).get("segment", "Equity")

        for cur_date in event_dates:
            day_snap = date_snaps.get(cur_date, {})
            is_lat   = (cur_date == latest_new)
            tgt      = latest_events if is_lat else historical_events

            if code not in day_snap:
                # Stock absent from this block → it exited here
                if prev_w > 0.01:
                    tgt.append({"nseCode": code, "securityName": last_sn, "segment": last_seg,
                                "eventType": "sell", "date": cur_date, "delta": round(prev_w, 4),
                                "newWeight": 0.0, "isSeriesReset": True})
                prev_w = 0.0
                continue

            s       = day_snap[code]
            new_w   = s["weight"]
            last_sn  = s["securityName"]
            last_seg = s["segment"]

            if prev_w < 0.01:
                tgt.append({"nseCode": code, "securityName": last_sn, "segment": last_seg,
                            "eventType": "buy", "date": cur_date, "delta": round(new_w, 4),
                            "newWeight": round(new_w, 4), "isSeriesReset": False})
            elif new_w > prev_w + 0.01:
                tgt.append({"nseCode": code, "securityName": last_sn, "segment": last_seg,
                            "eventType": "buy", "date": cur_date, "delta": round(new_w - prev_w, 4),
                            "newWeight": round(new_w, 4), "isSeriesReset": False})
            elif new_w < prev_w - 0.01:
                tgt.append({"nseCode": code, "securityName": last_sn, "segment": last_seg,
                            "eventType": "sell", "date": cur_date, "delta": round(prev_w - new_w, 4),
                            "newWeight": round(new_w, 4), "isSeriesReset": False})

            prev_w = new_w

    # ── Slide 2: compare latest block vs the IMMEDIATELY PRECEDING block ──
    # Use event_dates (not new_dates) so that zombie dates — old dates missing from
    # rebalance_history due to the previous current_codes filter — don't corrupt the
    # baseline. If event_dates has >1 entry we compare the latest against the
    # second-to-last event date (correct for a multi-date first upload). If only one
    # event date exists, compare against existing_weights (the confirmed prior state).
    if len(event_dates) > 1:
        prev_snap_s2 = date_snaps.get(event_dates[-2], {})
        prev_w_s2    = {c: s["weight"] for c, s in prev_snap_s2.items()}
    else:
        prev_snap_s2 = {}
        prev_w_s2    = existing_weights

    wholly_sold_s2 = set(prev_w_s2.keys()) - current_codes

    slide2: list[dict] = []
    for s in date_stock_map.get(latest_new, []):
        code   = s["nseCode"]
        prev_w = prev_w_s2.get(code, 0.0)
        new_w  = s["weight"]
        if prev_w < 0.01:
            ut, delta = "New Addition", round(new_w, 4)
        elif new_w > prev_w + 0.01:
            ut, delta = "Partial Add", round(new_w - prev_w, 4)
        elif new_w < prev_w - 0.01:
            ut, delta = "Partial Sell", round(prev_w - new_w, 4)
        else:
            ut, delta = "No Change", 0.0
        chg = (f"+{round(new_w - prev_w, 2)}%" if new_w > prev_w + 0.01
               else f"-{round(prev_w - new_w, 2)}%" if prev_w > new_w + 0.01
               else "No change")
        slide2.append({"nseCode": code, "stockName": s["securityName"], "segment": s["segment"],
                        "prevWeight": round(prev_w, 2), "newWeight": round(new_w, 2),
                        "delta": delta, "change": chg, "updateType": ut, "eventDate": latest_new})

    for code in wholly_sold_s2:
        prev_w     = prev_w_s2.get(code, 0.0)
        snap_entry = prev_snap_s2.get(code, {})
        sn  = snap_entry.get("securityName") or basket_bp_curr.get(code, {}).get("securityName", code)
        seg = snap_entry.get("segment")      or basket_bp_curr.get(code, {}).get("segment", "Equity")
        slide2.append({"nseCode": code, "stockName": sn, "segment": seg,
                        "prevWeight": round(prev_w, 2), "newWeight": 0.0,
                        "delta": round(prev_w, 4), "change": "Removed", "updateType": "Wholly Sell",
                        "eventDate": latest_new})

    # Slide 1: date discrepancies — compare Excel date vs existing last event date (±7 days)
    slide1: list[dict] = []
    def _last_event_date(det: dict, evt_type: str) -> str | None:
        field = "buyEvents" if evt_type == "buy" else "sellEvents"
        lines = [l.strip() for l in (det.get(field) or "").strip().split("\n") if " * " in l]
        if not lines:
            return None
        try:
            return lines[-1].split(" * ")[0].strip()
        except Exception:
            return None

    for evt in latest_events:
        code      = evt["nseCode"]
        excel_dt  = evt["date"]
        exist_dt  = _last_event_date(basket_bp_curr.get(code, {}), evt["eventType"])
        if not exist_dt or exist_dt == excel_dt:
            continue
        try:
            diff = abs((datetime.strptime(exist_dt, "%d %b %Y") -
                        datetime.strptime(excel_dt,  "%d %b %Y")).days)
        except Exception:
            continue
        if 0 < diff <= 7:
            slide1.append({"nseCode": code, "stockName": evt["securityName"],
                           "eventType": "Buy" if evt["eventType"] == "buy" else "Sell",
                           "existingDate": exist_dt, "newDate": excel_dt, "diffDays": diff})

    # History entries (for rebalance_history.json on confirm)
    # Save ALL stocks at every new date so that every date is tracked in existing_dates
    # on the next upload — prevents zombie dates from reappearing as "new".
    history_entries: list[dict] = []
    for d in new_dates:
        for s in date_stock_map.get(d, []):
            history_entries.append({"nseCode": s["nseCode"], "securityName": s["securityName"],
                                    "segment": s["segment"], "weight": round(s["weight"], 4), "date": d})

    return {
        "duplicate": False,
        "basketKey": basket,
        "basket": BASKET_DISPLAY_NAMES[basket],
        "newDates": new_dates,
        "latestDate": latest_new,
        "slide1": slide1,
        "slide2": slide2,
        "historicalEvents": historical_events,
        "historyEntries": history_entries,
    }


@router.post("/api/confirm-rebalance")
async def confirm_rebalance(
    background_tasks: BackgroundTasks,
    request: Request,
):
    """Apply confirmed rebalance changes from the 2-slide preview modal."""
    admin_email = _require_admin(request)
    body          = await request.json()
    basket        = body.get("basket", "")
    latest_date   = body.get("latestDate", "")
    slide2        = body.get("slide2", [])        # confirmed (possibly edited) latest-block rows
    hist_events   = body.get("historicalEvents", [])   # pre-computed older-block events
    hist_entries  = body.get("historyEntries", [])     # rows for rebalance_history.json

    if basket not in BASKET_DISPLAY_NAMES:
        raise HTTPException(400, f"Unknown basket: {basket}")
    if not slide2:
        raise HTTPException(400, "No confirmed changes provided.")

    _auto_save_rollback()
    _push_undo_snapshot(basket, f"before rebalance {latest_date}")

    bp_data   = _load_buy_price_data()
    basket_bp = bp_data.setdefault(basket, {})

    portfolios  = _load_portfolios()
    basket_stks = portfolios.get(basket, [])
    stk_map     = {s["nseCode"]: s for s in basket_stks}

    # Maps event_date → lists of buy/sell codes (for OHLC fetch)
    new_buys_by_date:  dict[str, list[str]] = {}
    new_sells_by_date: dict[str, list[str]] = {}

    def _apply_event(code: str, evt_type: str, date: str, delta: float,
                     sec_name: str = "", segment: str = "Equity",
                     series_reset: bool = False):
        _add_event(basket_bp, code, f"{evt_type}Events", date, delta)
        det = basket_bp[code]
        if not det.get("securityName") and sec_name:
            det["securityName"] = sec_name
        if not det.get("segment") and segment:
            det["segment"] = segment
        if series_reset:
            det["prevBuyEvents"]  = det.get("buyEvents",  "")
            det["prevSellEvents"] = det.get("sellEvents", "")
            det["buyEvents"]  = ""
            det["sellEvents"] = ""

    # 1. Apply all historical events to buy_price_data only (no portfolio changes here)
    for evt in hist_events:
        _apply_event(evt["nseCode"], evt["eventType"], evt["date"], evt["delta"],
                     evt.get("securityName", ""), evt.get("segment", "Equity"),
                     evt.get("isSeriesReset", False))
        if evt["eventType"] == "buy" and evt.get("newWeight", 0) > 0:
            new_buys_by_date.setdefault(evt["date"], []).append(evt["nseCode"])
        elif evt["eventType"] == "sell":
            new_sells_by_date.setdefault(evt["date"], []).append(evt["nseCode"])

    # 2. Apply latest-block buy/sell events to buy_price_data
    for row in slide2:
        code       = row["nseCode"]
        ut         = row["updateType"]
        event_date = row.get("eventDate", latest_date)
        new_weight = float(row.get("newWeight", 0))
        prev_weight = float(row.get("prevWeight", 0))
        delta      = float(row.get("delta", abs(new_weight - prev_weight)))
        sec_name   = row.get("stockName", "")
        segment    = row.get("segment", "Equity")

        if ut == "No Change":
            pass  # no event, handled in portfolio rebuild below
        elif ut in ("New Addition", "Partial Add"):
            _apply_event(code, "buy", event_date, delta, sec_name, segment)
            if ut == "New Addition":
                new_buys_by_date.setdefault(event_date, []).append(code)
        elif ut == "Partial Sell":
            _apply_event(code, "sell", event_date, delta, sec_name, segment)
            new_sells_by_date.setdefault(event_date, []).append(code)
        elif ut == "Wholly Sell":
            _apply_event(code, "sell", event_date, delta, sec_name, segment, series_reset=True)
            new_sells_by_date.setdefault(event_date, []).append(code)

    # 3. Rebuild portfolio from slide2 — the latest block is the authoritative composition.
    # Preserve all existing entry fields (buyPrice, live data, etc.) for stocks already in portfolio.
    # Stocks absent from slide2 (exited in historical blocks) are naturally excluded.
    new_basket_stks: list[dict] = []
    for row in slide2:
        code  = row["nseCode"]
        ut    = row.get("updateType", "")
        new_w = float(row.get("newWeight", 0))
        if ut == "Wholly Sell" or new_w <= 0:
            continue
        alloc          = round(new_w / 100, 6)
        existing_entry = stk_map.get(code)
        if existing_entry:
            new_basket_stks.append({**existing_entry, "allocation": alloc})
        else:
            new_basket_stks.append({"nseCode": code, "allocation": alloc, "buyPrice": None})

    # 4. Persist rebalance history
    rh = _load_rebalance_history()
    existing_dates = {e.get("date", "").strip() for e in rh.get(basket, [])}
    for entry in hist_entries:
        if entry.get("date", "") not in existing_dates:
            rh.setdefault(basket, []).append(entry)
    _save_rebalance_history(rh)

    # 5. Rebuild sold records from the now-updated event log (authoritative)
    old_sold    = portfolios.get(f"{basket}_sold", [])
    sold_stocks = _rebuild_sold_from_bp(basket_bp, old_sold)

    # 6. Persist portfolios (active + sold) and buy_price_data
    portfolios[basket] = new_basket_stks
    portfolios[f"{basket}_sold"] = sold_stocks
    _save_portfolios(portfolios)
    _save_buy_price_data(bp_data)

    # 7. Background tasks — fetch OHLC for both buy and sell events
    all_evt_dates = sorted(set(list(new_buys_by_date.keys()) + list(new_sells_by_date.keys())))
    for evt_date in all_evt_dates:
        b_codes = new_buys_by_date.get(evt_date, [])
        s_codes = new_sells_by_date.get(evt_date, [])
        if b_codes or s_codes:
            background_tasks.add_task(_fetch_rebalance_prices, basket, evt_date, b_codes, s_codes)
    background_tasks.add_task(_recalc_basket_buy_prices, basket)
    # Backfill fills any missing sell OHLC across all baskets and regenerates gains as final step
    background_tasks.add_task(_backfill_all_sell_ohlc_bg)

    _log_activity("rebalance_upload", admin_email, {
        "basket": basket,
        "basketLabel": BASKET_DISPLAY_NAMES.get(basket, basket),
        "date": latest_date,
        "stocksProcessed": len(slide2),
    })

    return {
        "ok": True,
        "basket": BASKET_DISPLAY_NAMES[basket],
        "date": latest_date,
        "stocksProcessed": len(slide2),
    }


@router.get("/api/activity-log")
async def get_activity_log(request: Request):
    """Return recent rebalance activity log — admin only."""
    _require_admin(request)
    log = json.loads(_ACTIVITY_LOG_FILE.read_text()) if _ACTIVITY_LOG_FILE.exists() else []
    return log


def _parse_excel_date(val) -> str | None:
    """Convert openpyxl cell value (datetime, date, or string) → 'DD Mon YYYY'."""
    if val is None:
        return None
    if isinstance(val, (datetime,)):
        return val.strftime("%d %b %Y")
    try:
        from datetime import date as _date
        if isinstance(val, _date):
            return val.strftime("%d %b %Y")
    except Exception:
        pass
    s = str(val).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d %b %Y", "%d-%b-%Y",
                "%d %B %Y", "%d-%B-%Y", "%d/%b/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d %b %Y")
        except ValueError:
            pass
    return None


def _resolve_nse_code(stock_name: str, basket_key: str, bp_data: dict) -> str:
    """Best-effort mapping of a stock name → NSE code.
    Priority: exact code match in basket → securityName match in basket → use as-is."""
    name = stock_name.strip()
    basket_bp = bp_data.get(basket_key, {})

    # 1. Exact NSE code match (case-insensitive)
    upper = name.upper()
    if upper in basket_bp:
        return upper
    for code in basket_bp:
        if code.upper() == upper:
            return code

    # 2. securityName match (case-insensitive)
    lower = name.lower()
    for code, det in basket_bp.items():
        sn = (det.get("securityName") or "").lower()
        if sn and (sn == lower or sn.startswith(lower) or lower.startswith(sn)):
            return code

    # 3. Fall back: treat value as NSE code
    return upper
