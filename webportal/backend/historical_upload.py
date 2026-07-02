"""Historical rebalance-constituents Excel import, plus rebuilding the sold-
stocks list from buy-price data (recovery tool after data corrections)."""
import io

import openpyxl
from fastapi import APIRouter, BackgroundTasks, File, Form, HTTPException, UploadFile

from buy_price_gains import (
    _date_to_ts, _add_event,
    _rebuild_sold_from_bp, _recalc_basket_buy_prices, _refresh_gains_file,
)
from persistence import (
    BASKET_DISPLAY_NAMES, _auto_save_rollback,
    _load_portfolios, _save_portfolios,
    _load_buy_price_data, _save_buy_price_data,
    _load_rebalance_history, _save_rebalance_history,
)
from rebalance import _parse_excel_date, _resolve_nse_code

router = APIRouter()

@router.post("/api/upload-historical-excel")
async def upload_historical_excel(
    background_tasks: BackgroundTasks,
    basket: str = Form(...),
    file: UploadFile = File(...),
):
    """Upload an Excel workbook whose 'Historical Constituents' sheet (col A: Date,
    col B: Stock Name, col C: Weight %) contains rebalance history.
    Only dates AFTER the last stored rebalance date are processed."""

    if basket not in BASKET_DISPLAY_NAMES:
        raise HTTPException(status_code=400, detail=f"Unknown basket: {basket}")

    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot open Excel file: {e}")

    # Locate the sheet
    sheet = None
    for name in wb.sheetnames:
        if "historical constituent" in name.lower():
            sheet = wb[name]
            break
    if sheet is None and len(wb.sheetnames) >= 2:
        sheet = wb.worksheets[1]   # fall back to sheet 2
    if sheet is None:
        raise HTTPException(
            status_code=400,
            detail="Sheet 'Historical Constituents' not found. "
                   "Expected sheet 2 or a sheet named 'Historical Constituents'.",
        )

    # Parse rows → { date_str: { stock_name: weight } }
    by_date: dict[str, dict[str, float]] = {}
    skipped = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        date_val, stock_val, weight_val = row[0], row[1], row[2]
        if not date_val or not stock_val:
            continue
        date_str = _parse_excel_date(date_val)
        if not date_str:
            skipped += 1
            continue
        try:
            weight = float(weight_val or 0)
        except (TypeError, ValueError):
            weight = 0.0
        stock_name = str(stock_val).strip()
        if not stock_name:
            continue
        by_date.setdefault(date_str, {})[stock_name] = round(weight, 6)

    if not by_date:
        raise HTTPException(status_code=400, detail="No valid data rows found in the sheet.")

    # Sort dates chronologically
    sorted_dates = sorted(by_date.keys(), key=_date_to_ts)

    # Find the last stored rebalance date for this basket
    rh = _load_rebalance_history()
    existing_entries = rh.get(basket, [])
    stored_dates = {e["date"] for e in existing_entries}
    last_stored_ts = max((_date_to_ts(d) for d in stored_dates), default=0)

    # Determine the previous snapshot at last_stored_ts
    # (latest snapshot from existing history whose date == max stored date)
    if stored_dates:
        last_stored_date = max(stored_dates, key=_date_to_ts)
        prev_snap: dict[str, float] = {
            e["nseCode"]: e["weight"]
            for e in existing_entries
            if e["date"] == last_stored_date
        }
    else:
        prev_snap = {}

    # Load supporting data
    bp_data    = _load_buy_price_data()
    portfolios = _load_portfolios()
    basket_bp  = bp_data.setdefault(basket, {})
    basket_stks = portfolios.setdefault(basket, [])
    stk_map    = {s["nseCode"]: s for s in basket_stks}
    sold       = portfolios.setdefault(f"{basket}_sold", [])

    # Process only dates strictly after last_stored_ts
    new_dates = [d for d in sorted_dates if _date_to_ts(d) > last_stored_ts]

    if not new_dates:
        return {
            "ok": True,
            "message": "No new rebalance dates found after the last stored date "
                       f"({max(stored_dates, key=_date_to_ts) if stored_dates else 'none'}).",
            "newDatesProcessed": 0,
        }

    summary: list[dict] = []
    rh_new_entries: list[dict] = []
    bp_changed = False

    for date_str in new_dates:
        if date_str in stored_dates:
            continue   # already stored — skip

        curr_snap_raw: dict[str, float] = by_date[date_str]
        # Resolve stock names → NSE codes
        curr_snap: dict[str, float] = {}
        for stock_name, weight in curr_snap_raw.items():
            code = _resolve_nse_code(stock_name, basket, bp_data)
            curr_snap[code] = weight

        changes: list[dict] = []

        # Stocks in current snapshot
        for code, weight in curr_snap.items():
            prev_weight = prev_snap.get(code, 0.0)
            delta = round(weight - prev_weight, 6)

            if prev_weight == 0.0 and weight > 0:
                action = "Fresh Addition"
                # Buy event
                _add_event(basket_bp, code, "buyEvents", date_str, weight)
                bp_changed = True
                # Add to portfolio if absent
                if code not in stk_map:
                    entry = {"nseCode": code, "allocation": round(weight / 100, 6), "buyPrice": None}
                    basket_stks.append(entry)
                    stk_map[code] = entry
                else:
                    stk_map[code]["allocation"] = round(
                        stk_map[code].get("allocation", 0) + weight / 100, 6)
            elif delta > 0.001:
                action = "Addition"
                _add_event(basket_bp, code, "buyEvents", date_str, delta)
                bp_changed = True
                if code in stk_map:
                    stk_map[code]["allocation"] = round(weight / 100, 6)

            elif delta < -0.001:
                sell_qty = abs(delta)
                action = "Partial Sell" if weight > 0.001 else "Full Removal"
                _add_event(basket_bp, code, "sellEvents", date_str, sell_qty)
                bp_changed = True

                if weight <= 0.001:
                    # Remove from active portfolio → sold list
                    sold.append({
                        "nseCode": code,
                        "securityName": (basket_bp.get(code) or {}).get("securityName", ""),
                        "date": date_str,
                        "action": "Wholly Sold",
                        "weightSold": round(prev_weight, 2),
                        "buyPrice": stk_map.get(code, {}).get("buyPrice"),
                        "sellPrice": None,
                    })
                    basket_stks = [s for s in basket_stks if s["nseCode"] != code]
                    stk_map.pop(code, None)
                else:
                    if code in stk_map:
                        stk_map[code]["allocation"] = round(weight / 100, 6)
            else:
                action = "Unchanged"

            changes.append({"code": code, "action": action,
                             "prev": prev_weight, "curr": weight})

        # Stocks present in prev but absent in curr → fully removed
        for code, prev_weight in prev_snap.items():
            if code not in curr_snap and prev_weight > 0.001:
                _add_event(basket_bp, code, "sellEvents", date_str, prev_weight)
                bp_changed = True
                basket_stks = [s for s in basket_stks if s["nseCode"] != code]
                stk_map.pop(code, None)
                changes.append({"code": code, "action": "Full Removal (absent from new snapshot)",
                                 "prev": prev_weight, "curr": 0})

        # Append to rebalance history
        for code, weight in curr_snap.items():
            sn = (basket_bp.get(code) or {}).get("securityName", "")
            rh_new_entries.append({
                "date": date_str, "nseCode": code,
                "securityName": sn, "segment": "", "weight": weight,
            })

        non_unchanged = [c for c in changes if c["action"] != "Unchanged"]
        summary.append({"date": date_str, "changes": non_unchanged,
                        "total": len(curr_snap), "changed": len(non_unchanged)})
        prev_snap = curr_snap   # roll forward

    # Persist
    rh.setdefault(basket, []).extend(rh_new_entries)
    _save_rebalance_history(rh)

    if bp_changed:
        bp_data[basket] = basket_bp
        _save_buy_price_data(bp_data)

    # Always rebuild sold records from the updated event log
    portfolios[basket] = basket_stks
    portfolios[f"{basket}_sold"] = _rebuild_sold_from_bp(basket_bp, sold)
    _save_portfolios(portfolios)

    # Background: recalc buy prices + refresh gains
    background_tasks.add_task(_recalc_basket_buy_prices, basket)
    background_tasks.add_task(_refresh_gains_file)

    return {
        "ok": True,
        "basket": BASKET_DISPLAY_NAMES[basket],
        "newDatesProcessed": len(new_dates),
        "skippedRows": skipped,
        "summary": summary,
    }


@router.post("/api/rebuild-sold/{basket}")
async def rebuild_sold_endpoint(basket: str, background_tasks: BackgroundTasks):
    """Rebuild sold-stock records from buy/sell event log. Fixes wrong weights, actions,
    sell prices, and duplicates caused by earlier code paths."""
    if basket not in BASKET_DISPLAY_NAMES:
        raise HTTPException(400, f"Unknown basket: {basket}")
    _auto_save_rollback()
    bp_data    = _load_buy_price_data()
    basket_bp  = bp_data.get(basket, {})
    portfolios = _load_portfolios()
    old_sold   = portfolios.get(f"{basket}_sold", [])
    new_sold   = _rebuild_sold_from_bp(basket_bp, old_sold)
    portfolios[f"{basket}_sold"] = new_sold
    _save_portfolios(portfolios)
    background_tasks.add_task(_recalc_basket_buy_prices, basket)
    background_tasks.add_task(_refresh_gains_file)
    return {"ok": True, "basket": BASKET_DISPLAY_NAMES[basket], "recordCount": len(new_sold)}


